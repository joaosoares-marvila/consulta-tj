from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from loguru import logger
from pathlib import Path
from datetime import datetime
import os
import traceback

class Planilha:
    def __init__(self, caminho_arquivo: str, nome_planilha: str, cabecalho: list[str]) -> None:
        """
        Inicializa a instância da classe Planilha.
        """

        # Planilha de processos
        self.caminho_arquivo = caminho_arquivo
        self.nome_planilha = nome_planilha
        self.cabecalho = cabecalho
        
        self.__obtem_wb()
        self.__obtem_ws()


    
    def __obtem_wb(self) -> bool:

        if os.path.exists(self.caminho_arquivo):
            self.wb = load_workbook(filename=self.caminho_arquivo)
        else:
            self.wb = Workbook()
        
        self.wb.save(filename=self.caminho_arquivo)

    def __obtem_ws(self) -> None:
        
        if self.nome_planilha in self.wb.sheetnames:
            self.ws = self.wb[self.nome_planilha]
        else:
            self.ws = self.wb.create_sheet(title=self.nome_planilha, index=0)
            self.__define_cabecalho()

    def __define_cabecalho(self) -> None:

        i_coluna = 1

        for nome_coluna in self.cabecalho:

            nome_coluna = str(nome_coluna).replace('_', ' ').upper()
            
            # Define o cabeçalho (linha 1) da coluna com o seu respectivo valor
            celula_coluna = self.ws.cell(row=1, column=i_coluna, value=nome_coluna)

            # Aplica estilos ao cabeçalho
            celula_coluna.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
            celula_coluna.fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            celula_coluna.alignment = Alignment(horizontal='center')
            celula_coluna.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            i_coluna += 1
        
        self.wb.save(filename=self.caminho_arquivo)
        
    def __obter_proxima_linha_vazia(self) -> int:
        """
        Obtém o número da próxima linha vazia na planilha.

        Returns:
            int: Número da próxima linha vazia.
        """
        for row in range(2, self.ws.max_row + 2):  # Começa da segunda linha
            if not self.ws.cell(row=row, column=1).value:  # Verifica a primeira coluna
                return row
            
    def __obtem_indice_coluna(self, coluna_desejada: str) -> int:
        
        # Encontrar o índice da coluna com o cabeçalho desejado
        indice_coluna = None
        for col in range(1, self.ws.max_column + 1):
            if self.ws.cell(row=1, column=col).value == coluna_desejada.upper():
                indice_coluna = col
                break
    
        return indice_coluna

    def adiciona_dado(self, dados: dict) -> None:

        try:
            
            linha = self.__obter_proxima_linha_vazia()

            for nome_coluna, valor in dados.items():
                
                nome_coluna = str(nome_coluna).replace('_', ' ').upper()
                coluna = self.__obtem_indice_coluna(coluna_desejada=nome_coluna)

                # Preenche o campo referente à coluna e linha atual com o seu respectivo dado
                celula_dado = self.ws.cell(row=linha, column=coluna, value=valor)

                # Aplica estilos diretamente à célula de dado
                celula_dado.font = Font(name='Arial', size=11, bold=False)
                celula_dado.alignment = Alignment(horizontal='left')
                celula_dado.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

                self.wb.save(filename=self.caminho_arquivo)

        except:
            logger.erro(traceback.format_exc())
    def obtem_processos_planilha(self) -> list:

        coluna = self.__obtem_indice_coluna(coluna_desejada='processo')
        processos = [self.ws.cell(row=row, column=coluna).value for row in range(2, self.ws.max_row + 1)]

        return processos
    
    def obtem_foros(self) -> list:

        coluna = self.__obtem_indice_coluna(coluna_desejada='foro')
        processos = [self.ws.cell(row=row, column=coluna).value for row in range(2, self.ws.max_row + 1)]

        return processos

    def salvar(self) -> None:
        self.wb.save(filename=self.caminho_arquivo)

    def obtem_url_processos(self)-> list[dict]:
        
        processos = []

        # Iterar sobre as linhas da planilha
        for row in self.ws.iter_rows():

            processo = row[0].value
            url = row[1].value
            
            processos.append({
                'processo': processo,
                'url': url
            })

        return processos
    
    def obter_foros_pendentes(self) -> list[dict]:

        # Lista para armazenar os valores da coluna 2 onde a coluna 3 está vazia
        foros_pendentes = []

        # Iterar sobre as linhas da planilha
        for row in self.ws.iter_rows():

            if row[2].value is None:
                dados_foro = {
                    'foro': row[0].value,
                    'url_pesquisa': row[1].value
                }
                foros_pendentes.append(dados_foro)

        return foros_pendentes
    
    def preencher_sitaucao_foro(self, foro: str, situacao: str) -> None:

        for row in self.ws.iter_rows():
            if row[0].value == foro:
                row[2].value = situacao
                row[3].value = datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
                break
        
        self.wb.save(filename=self.caminho_arquivo)

    def define_horario_verificacao_processo(self, numero_processo)->None:

        for row in self.ws.iter_rows():
            if row[0].value == numero_processo:
                if row[12].value is None:
                    row[12].value = datetime.now().strftime("%m/%d/%Y, %H:%M:%S")
                    self.wb.save(filename=self.caminho_arquivo)
                    break
        


if __name__ == '__main__':
    ...